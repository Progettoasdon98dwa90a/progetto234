from datetime import datetime, timedelta
from io import BytesIO

from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl.reader.excel import load_workbook

from api.models import Branch, Import, Employee


def getHistory(request, year, branch_id):

    if request.method == 'GET':
        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Branch not found'})

        try:
            year = int(year)
            year = str(year)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid year'})

        imports_objs_qs = Import.objects.filter(branch=branch)
        imports_objs_qs_filtered = Import.objects.none()

        # Filter for all import_obj containing the requested year
        for import_obj in imports_objs_qs:
            if import_obj.import_date.__contains__(year):
                imports_objs_qs_filtered |= Import.objects.filter(import_date=import_obj.import_date)

        # Build dictionary of all days in the selected year
        start_date = datetime(int(year), 1, 1)
        end_date = datetime(int(year), 12, 31)
        all_days_in_year = [
            (start_date + timedelta(days=i)).strftime('%Y-%m-%d')
            for i in range((end_date - start_date).days + 1)
        ]

        imports_report_mapped = {}

        # Mark days with import as 1
        for import_obj in imports_objs_qs_filtered:
            imports_report_mapped[import_obj.import_date] = 1
            if import_obj.import_date in all_days_in_year:
                all_days_in_year.remove(import_obj.import_date)

        # Remaining days get 0 for no import
        for day in all_days_in_year:
            imports_report_mapped[day] = 0

        # Now mark FUTURE days as -1
        today_str = datetime.now().strftime('%Y-%m-%d')
        today_date = datetime.strptime(today_str, '%Y-%m-%d').date()

        for day_str in imports_report_mapped.keys():
            day_date = datetime.strptime(day_str, '%Y-%m-%d').date()
            if day_date > today_date:
                imports_report_mapped[day_str] = -1

        # Sort dictionary by date
        imports_report_mapped = dict(sorted(imports_report_mapped.items()))

        context = {
            'imports': imports_report_mapped,
            'branch': branch.id,
            'year': year,
        }

        return JsonResponse(context)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def uploadImportData(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        selected_branch = request.POST.get('branchSelect')
        selected_type = request.POST.get('typeSelect')

        # 0 = sales_data, 1 = counter_data
        if selected_type == 0:
            selected_type = "sales_data"
        elif selected_type == 1:
            selected_type = "counter_data"

        try:
            selected_branch = int(selected_branch)
        except:
            return JsonResponse({"status": "error", "errors": ["Internal"]}, status=200)

        if not selected_branch:
            return JsonResponse({"status": "error", "errors": ["no branch selected"]}, status=200)

        data_dict = dict

        branch_obj = Branch.objects.get(id=selected_branch)

        if branch_obj.get_brand() == "equivalenza":
            if selected_type == "counter_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                        date = row[0]  # assuming the date is in the second column
                        record = {
                            "(Ing) Ingressi": row[1] or 0,
                            "(Est) Traffico Esterno": row[2] or 0,
                            "(TA) Tasso di Attrazione": row[3] or 0,
                        }

                        # mar-19.11.24 to YYYY/MM/DD
                        date = date.split("-")[1]
                        date = datetime.strptime(date, "%d.%m.%y").strftime("%Y-%m-%d")

                        # If the date is already a key, overwrite the current record
                        if date in data_dict:

                            data_dict[date] = record
                        else:
                            # Otherwise, create a new list with the record
                            data_dict[date] = [record]

                        data_dict[date] = [record]

                    for date, records in data_dict.items():
                        if Import.objects.filter(branch=branch_obj, import_date=date,
                                                 import_type=selected_type).exists():
                            errors.append(f"Collision on {date}")
                            continue

                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    import_bulk_create_list = []
                    for date, data in data_dict.items():
                        i = Import(import_date=date, data=data, branch=branch_obj, import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)

                    return JsonResponse({"status": "success"}, status=200)

            ### START CONVERTING DATA
            elif selected_type == "sales_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                        date = row[1]  # assuming the date is in the second column
                        record = {
                            "Dipendente": row[0],
                            "Qta. Vend.": row[2],
                            "Sco.": row[3],
                            "Importo": row[4],
                            "Sco. Medio": row[5],
                            "Qta Media": row[6]
                        }

                        # If the date is already a key, append the new record to its list
                        if date in data_dict:
                            data_dict[date].append(record)
                        else:
                            # Otherwise, create a new list with the record
                            data_dict[date] = [record]

                    import_qs = Import.objects.none()

                    errors = []

                    for date, records in data_dict.items():
                        employees_day_id_list = []
                        date = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
                        for record in records:
                            try:
                                employees_day_id_list.append(record['Dipendente'])
                            except:
                                errors.append(f"Employee {record['Dipendente']} not found")
                                continue

                        employees_day_qs = Employee.objects.filter(id__in=employees_day_id_list)

                        if employees_day_qs.count() != len(employees_day_id_list):
                            errors.append(f"Employees on {date} not found")
                            continue

                        branch_check_value = list(employees_day_qs.values_list('branch', flat=True))
                        normalized = list(dict.fromkeys(branch_check_value))
                        if len(normalized) != 1:
                            errors.append(f"Branch on {date} from different branch")
                            continue

                        if normalized[0] != selected_branch:
                            errors.append(f"Branch on {date} from different branch")
                            continue

                        if Import.objects.filter(branch=branch_obj, import_date=date):
                            errors.append(f"Collision on {date}")
                            continue

                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    import_bulk_create_list = []
                    for date, data in data_dict.items():
                        date_obj = datetime.strptime(date, '%d/%m/%Y')
                        date_str = date_obj.strftime('%Y-%m-%d')
                        i = Import(import_date=date_str, data=data, branch=branch_obj,
                                   import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)

                return JsonResponse({"status": "success", "errors": []}, status=200)

        elif branch_obj.get_brand() == "original":
            print("original")

            if selected_type == "counter_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row

                        date = row[0]  # assuming the date is in the second column
                        record = {
                            "(Ing) Ingressi": row[1],
                            "(Est) Traffico Esterno": row[2],
                            "(TA) Tasso di Attrazione": row[3],
                        }

                        data_dict[date] = [record]

            ### START CONVERTING DATA
            elif selected_type == "sales_data":

                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row

                        if row[1] in ["Totale generale", "Totali "]:
                            continue

                        date = row[1]  # assuming the date is in the second column
                        record = {
                            "Qta. Vend.": row[2],
                            "Sco.": row[3],
                            "Importo": row[4],
                            "Sco. Medio": row[5],
                            "Qta Media": row[6]
                        }
                        if Import.objects.filter(branch=branch_obj, import_date=date):
                            errors.append(f"Collision on {date}")
                            continue

                        data_dict[date] = [record]

                    import_qs = Import.objects.none()

                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    data_dict_formatted = {}

                    data_dict = {k: v for k, v in data_dict.items() if k is not None}

                    for k, v in data_dict.items():
                        unformatted = k
                        date_strp = datetime.strptime(unformatted, '%d/%m/%Y')
                        date_str = date_strp.strftime('%Y-%m-%d')
                        data_dict_formatted[date_str] = v

                    import_bulk_create_list = []
                    for date, data in data_dict_formatted.items():
                        i = Import(import_date=date, data=data, branch=branch_obj, import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)

                return JsonResponse({"status": "success", "errors": []}, status=200)

        return JsonResponse({"status": "error", "errors": ["Internal"]}, status=200)
    return JsonResponse({"status": "error", "errors": ["Internal"]}, status=200)