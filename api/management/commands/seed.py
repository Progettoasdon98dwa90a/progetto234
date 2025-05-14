from datetime import datetime
from pathlib import Path
from django.core.management import BaseCommand
from django.http import JsonResponse
from openpyxl import load_workbook
from django.core.management import call_command


from api.models import Employee,Branch, Import, Schedule, Target

current_directory = Path(__file__).resolve().parent.parent.parent.parent


class Command(BaseCommand):
    help = "Assigns all weapons to users, without duplicates, until finished"

    def handle(self, *args, **options):
        # create first two branches
        selected_branch = 1

        branch_data = [
                {'name': 'Biella', 'extra_data': {'brand' : 'equivalenza'}},
                {'name': 'OM-Siderno', 'extra_data': {'brand' : 'original'}},
            ]

        for branch in branch_data:
            Branch.objects.create(**branch)
            print(f"Branch '{branch['name']}' created successfully.")


        branch_obj = Branch.objects.get(id=selected_branch)
        roles_data = {
                'Operatore': {

                },
            }


        employees_data = [{
                    'id': 1,
                    'first_name': 'Elisa',
                    'last_name': "1",
                    'branch': branch_obj,
                    'role': 0,  # Assuming you have a Role model
                    'max_hours_per_day': 8,
                    'max_hours_per_week': 40,
                    'max_hours_per_month': 160,
                    
                },
                {
                    'id': 2,
                    'first_name': 'Vanessa',
                    'last_name': "1",
                    'branch': branch_obj,
                    'role': 0,  # Assuming you have a Role model
                    'max_hours_per_day': 8,
                    'max_hours_per_week': 40,
                    'max_hours_per_month': 160,
                    
                },
                {
                    'id': 3,
                    'first_name': 'Vanessa',
                    'last_name': "Brocca",
                    'branch': branch_obj,
                    'role': 1,  # Assuming you have a Role model
                    'max_hours_per_day': 8,
                    'max_hours_per_week': 40,
                    'max_hours_per_month': 160,
                    
                },
                {
                    'id': 4,
                    'first_name': 'Venditore',
                    'last_name': "Christmas",
                    'branch': branch_obj,
                    'role': 1,  # Assuming you have a Role model
                    'max_hours_per_day': 8,
                    'max_hours_per_week': 40,
                    'max_hours_per_month': 160,
                    
                },

            ]
        
        for employee in employees_data:
            employee_obj = Employee.objects.create(**employee)
            print(f"Employee '{employee['first_name']}' created successfully.")


        files = ['Dati-Biella-2023.xlsx', 'Dati-Biella-2024.xlsx', 'Dati-Biella-2025.xlsx']
        # Specify the path to the local file
        for file_name in files:
            import_data_file = current_directory / 'utils_files' / file_name

            # Load the workbook from the local file
            workbook = load_workbook(filename=import_data_file)

            # Access the sheet
            sheet = workbook.active  # You can specify sheet name if needed

            # Initialize the dictionary to hold the data
            data_dict = {}

            # Iterate through the rows, skipping the header
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                date = row[1]  # assuming the date is in the second column
                record = {
                    "Dipendente": row[0],
                    "Qta. Vend.": row[2],
                    "Sco.": row[3],
                    "Importo": row[4],
                    "Sco. Medio": row[5],
                    "Qta Media": row[6]
                }

                # If the date is already a key, append the new record to its list
                if date in data_dict:
                    data_dict[date].append(record)
                else:
                    # Otherwise, create a new list with the record
                    data_dict[date] = [record]

            import_qs = Import.objects.none()

            errors = []
            selected_branch = 1
            selected_type = 'sales_data'

            import_bulk_create_list = []
            for date, data in data_dict.items():
                date_obj = datetime.strptime(date, '%d/%m/%Y')
                date_str = date_obj.strftime('%Y-%m-%d')
                i = Import(import_date=date_str, data=data, branch=branch_obj, import_type=selected_type)
                import_bulk_create_list.append(i)

            Import.objects.bulk_create(import_bulk_create_list)
            print(f"Import data created successfully for  {file_name}.")
        
        selected_branch = 1
        branch_obj = Branch.objects.get(id=selected_branch)

        # get all branch employees
        employees = list(Employee.objects.filter(branch=branch_obj).values_list('id', flat=True))

        # create some mock schedules
        s1 = Schedule.objects.create(start_date="2025-05-01",
                                    end_date="2025-05-31",
                                    branch=branch_obj,
                                    employees=employees,
                                    shift_data=[{"name": "M", "start": "9:00", "end": "13:00", "minEmployees": "1"},
                                                {"name": "P", "start": "13:00", "end": "17:00", "minEmployees": "2"},
                                                {"name": "C", "start": "17:00", "end": "20:00", "minEmployees": "1"}],
                                    free_days=[{"employee_id": 1, "free_days": ["2025-05-01", "2025-05-08", "2025-05-15", "2025-05-22"]},
                                               {"employee_id": 2, "free_days": ["2025-05-05", "2025-05-12", "2025-05-19", "2025-05-26"]},
                                               {"employee_id": 3, "free_days": ["2025-05-06", "2025-05-13", "2025-05-20", "2025-05-27"]}]
                                     )

        s2 = Schedule.objects.create(start_date="2025-06-01",
                                     end_date="2025-06-30",
                                     branch=branch_obj,
                                     employees=employees,
                                     shift_data=[{"name": "M", "start": "9:00", "end": "13:00", "minEmployees": "1"},
                                                 {"name": "P", "start": "13:00", "end": "17:00", "minEmployees": "2"},
                                                 {"name": "C", "start": "17:00", "end": "20:00", "minEmployees": "1"}],
                                     free_days=[{"employee_id": 1,
                                                 "free_days": ["2025-06-01", "2025-06-08", "2025-06-15", "2025-06-22"]},
                                                {"employee_id": 2,
                                                 "free_days": ["2025-06-05", "2025-06-12", "2025-06-19", "2025-06-26"]},
                                                {"employee_id": 3,
                                                 "free_days": ["2025-06-06", "2025-06-13", "2025-06-20", "2025-06-27"]}]
                                     )
        s1.save()
        s2.save()

        target = 300
        start_date = "2025-05-01"
        end_date = "2025-05-31"
        branch_obj_2 = Branch.objects.get(id=2)
        t1 = Target.objects.create(sales_target=target, start_date=start_date, end_date=end_date, branch=branch_obj)
        t2 = Target.objects.create(sales_target=target, start_date=start_date, end_date=end_date, branch=branch_obj_2)
        t1.save()
        t2.save()

        print(f"Target for {branch_obj.name} created successfully.")

        print("Schedule created successfully.")